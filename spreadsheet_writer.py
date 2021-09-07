from openpyxl import load_workbook

template_path = 'spreadsheet-templates/AP Chinese/Helper Sheet/Blank.xlsx'
wb = load_workbook(template_path)
print("opened workbook")
main_sheet = wb["Formatted"]
helper = wb["Helper"]

default_path = 'demo.xlsx'
# default_path = None


def set_output(path):
    global default_path
    default_path = path


def write_row(li, row, col):
    for col, val in enumerate(li, start=col):
        main_sheet.cell(row=row, column=col).value = val


def write_key(key):
    write_row(key, 2, 2)


def write_responses(responses):
    for i, response in enumerate(responses):
        main_sheet.cell(row=4 + i, column=1).value = response['name']
        write_row(response['selection'], 4 + i, 2)


def save():
    if not default_path:
        return
    wb.save(default_path)


if __name__ == '__main__':
    const_key = [[[1, 2, 3, 0, 1], [3, 1, 1, 1, 1], [3, 1, 0, 2, 3], [0, 0, 3, 1, 1, 3, 2, 2, 3, 0],
                  [3, 2, 1, 3, 0, 2, 3, -1, -1, -1]],
                 [[3, 1, 3, 2, 3, 1, 2, 0, 1, 1], [1, 3, 2, 3, 2, 0, 3, 1, 1, 2], [1, 0, 2, 3, 0, 2, 0, 2, 1, 2],
                  [3, 0, 2, 0, 3, 3, -1, -1, -1, -1]]]
    write_key([])
